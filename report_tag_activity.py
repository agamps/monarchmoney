import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
DEFAULT_TRANSACTIONS = Path("data/all_transactions.csv")
DEFAULT_OUTPUT = Path("data/tag_activity.xlsx")

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Consolas", bold=True, color="FFFFFF")
AMOUNT_FORMAT = "#,##0.00"
DATE_FORMAT = "mm/dd/yyyy"
TOTAL_LABEL = "Total"
BLANK_LABEL = "(blank)"


def optional_date_input(value: str | None) -> str | None:
    if value is None:
        return None

    text = value.strip()
    if text == "" or text.casefold() in {"auto", "default", "none"}:
        return None
    return text


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel workbook of tagged transactions."
    )
    parser.add_argument(
        "--transactions",
        type=Path,
        default=DEFAULT_TRANSACTIONS,
        help="Source transactions CSV. Defaults to data/all_transactions.csv.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Path for the generated Excel workbook.",
    )
    parser.add_argument(
        "--lookback-years",
        type=int,
        default=3,
        help=(
            "Number of calendar years to include, ending with the latest transaction "
            "year unless --end-date is provided. Defaults to 3."
        ),
    )
    parser.add_argument(
        "--start-date",
        type=optional_date_input,
        default=None,
        help="Optional inclusive start date. Overrides --lookback-years start.",
    )
    parser.add_argument(
        "--end-date",
        type=optional_date_input,
        default=None,
        help="Optional inclusive end date. Defaults to the latest transaction date.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> pd.DataFrame:
    last_error: UnicodeDecodeError | None = None

    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except UnicodeDecodeError as e:
            last_error = e

    assert last_error is not None
    raise ValueError(
        f"Could not decode {path} using supported encodings: {', '.join(CSV_ENCODINGS)}"
    ) from last_error


def require_columns(df: pd.DataFrame, path: Path, columns: list[str]) -> None:
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(missing)}")


def parse_amount(value: object) -> float:
    text = str(value or "").strip()
    if text == "":
        return 0.0

    is_parenthesized = text.startswith("(") and text.endswith(")")
    if is_parenthesized:
        text = text[1:-1].strip()

    text = text.replace(",", "").replace("$", "")
    amount = pd.to_numeric(text, errors="coerce")
    if pd.isna(amount):
        return 0.0

    amount = float(amount)
    return -abs(amount) if is_parenthesized else amount


def parse_dates(values: pd.Series) -> pd.Series:
    try:
        return pd.to_datetime(values, format="mixed", errors="coerce")
    except TypeError:
        return pd.to_datetime(values, errors="coerce")


def parse_date(value: str, label: str) -> pd.Timestamp:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"Could not parse {label}: {value}")
    return pd.Timestamp(parsed).normalize()


def clean_dimension(values: pd.Series) -> pd.Series:
    return (
        values.fillna("")
        .astype(str)
        .str.strip()
        .replace("", BLANK_LABEL)
    )


def split_tags(value: object) -> list[str]:
    tags: list[str] = []
    seen: set[str] = set()

    for part in str(value or "").split(","):
        tag = part.strip()
        if tag == "":
            continue

        key = tag.casefold()
        if key in seen:
            continue

        seen.add(key)
        tags.append(tag)

    return tags


def prepare_transactions(df: pd.DataFrame, path: Path) -> pd.DataFrame:
    require_columns(df, path, ["Transaction ID", "Date", "Amount", "Tags"])

    prepared = df.copy()
    prepared["Date"] = parse_dates(prepared["Date"])
    prepared["Amount"] = prepared["Amount"].map(parse_amount)
    prepared["Abs Amount"] = prepared["Amount"].abs()
    prepared = prepared.dropna(subset=["Date"])
    prepared["Year"] = prepared["Date"].dt.year.astype(int)

    for column in ["Account", "Merchant", "Plaid Name", "Category", "Notes"]:
        if column not in prepared.columns:
            prepared[column] = BLANK_LABEL
        prepared[column] = clean_dimension(prepared[column])

    for column in ["Hide From Reports", "Needs Review"]:
        if column not in prepared.columns:
            prepared[column] = ""
        prepared[column] = prepared[column].fillna("").astype(str).str.strip()

    prepared["Tags"] = prepared["Tags"].fillna("").astype(str).str.strip()
    return prepared


def date_window(
    df: pd.DataFrame,
    *,
    lookback_years: int,
    start_date: str | None,
    end_date: str | None,
) -> tuple[pd.Timestamp, pd.Timestamp]:
    if lookback_years < 1:
        raise ValueError("--lookback-years must be at least 1.")
    if df.empty:
        raise ValueError("No dated transaction rows were found.")

    end = (
        parse_date(end_date, "--end-date")
        if end_date is not None
        else pd.Timestamp(df["Date"].max()).normalize()
    )
    start = (
        parse_date(start_date, "--start-date")
        if start_date is not None
        else pd.Timestamp(year=end.year - lookback_years + 1, month=1, day=1)
    )

    if start > end:
        raise ValueError("--start-date must be on or before --end-date.")
    return start, end


def filter_date_window(
    df: pd.DataFrame,
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
) -> pd.DataFrame:
    dates = df["Date"].dt.normalize()
    return df.loc[(dates >= start_date) & (dates <= end_date)].copy()


def explode_tags(df: pd.DataFrame) -> pd.DataFrame:
    tagged = df.copy()
    tagged["Tag"] = tagged["Tags"].map(split_tags)
    tagged = tagged.explode("Tag")
    tagged["Tag"] = tagged["Tag"].fillna("").astype(str).str.strip()
    tagged = tagged[tagged["Tag"] != ""].copy()
    return tagged


def append_summary_total(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    total_row = {column: "" for column in df.columns}
    total_row[df.columns[0]] = TOTAL_LABEL

    for column in ["Transaction Count", "Net Amount", "Total Abs Amount"]:
        if column in df.columns:
            total_row[column] = df[column].sum()

    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def tag_summary(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Tag",
        "Transaction Count",
        "Net Amount",
        "Total Abs Amount",
        "Average Amount",
        "Smallest Amount",
        "Largest Amount",
        "First Date",
        "Last Date",
    ]
    if df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        df.groupby("Tag", dropna=False)
        .agg(
            **{
                "Transaction Count": ("Transaction ID", "count"),
                "Net Amount": ("Amount", "sum"),
                "Total Abs Amount": ("Abs Amount", "sum"),
                "Average Amount": ("Amount", "mean"),
                "Smallest Amount": ("Amount", "min"),
                "Largest Amount": ("Amount", "max"),
                "First Date": ("Date", "min"),
                "Last Date": ("Date", "max"),
            }
        )
        .reset_index()
    )
    return summary.sort_values(
        ["Total Abs Amount", "Transaction Count", "Tag"],
        ascending=[False, False, True],
        kind="stable",
    )


def summary_by_dimensions(df: pd.DataFrame, dimensions: list[str]) -> pd.DataFrame:
    columns = dimensions + ["Transaction Count", "Net Amount", "Total Abs Amount"]
    if df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        df.groupby(dimensions, dropna=False)
        .agg(
            **{
                "Transaction Count": ("Transaction ID", "count"),
                "Net Amount": ("Amount", "sum"),
                "Total Abs Amount": ("Abs Amount", "sum"),
            }
        )
        .reset_index()
    )
    return summary.sort_values(
        dimensions + ["Total Abs Amount"],
        ascending=[True] * len(dimensions) + [False],
        kind="stable",
    )


def tag_year_pivot(df: pd.DataFrame, years: list[int]) -> pd.DataFrame:
    year_labels = [str(year) for year in years]
    columns = ["Tag"] + year_labels + [TOTAL_LABEL]
    if df.empty:
        return pd.DataFrame(columns=columns)

    pivot = pd.pivot_table(
        df,
        index="Tag",
        columns="Year",
        values="Amount",
        aggfunc="sum",
        fill_value=0.0,
    )

    for year in years:
        if year not in pivot.columns:
            pivot[year] = 0.0

    pivot = pivot[years]
    pivot[TOTAL_LABEL] = pivot.sum(axis=1)
    pivot = pivot.reset_index()
    pivot.columns = [str(column) for column in pivot.columns]
    pivot = pivot.sort_values(TOTAL_LABEL, ascending=False, kind="stable")

    total_row = {column: "" for column in pivot.columns}
    total_row["Tag"] = TOTAL_LABEL
    for column in year_labels + [TOTAL_LABEL]:
        total_row[column] = pivot[column].sum()

    return pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)


def detail_columns(df: pd.DataFrame) -> list[str]:
    preferred_columns = [
        "Tag",
        "Transaction ID",
        "Date",
        "Year",
        "Account",
        "Merchant",
        "Plaid Name",
        "Amount",
        "Abs Amount",
        "Category",
        "Tags",
        "Notes",
        "Hide From Reports",
        "Needs Review",
    ]
    return [column for column in preferred_columns if column in df.columns]


def detail_by_amount(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=detail_columns(df))

    return df[detail_columns(df)].sort_values(
        ["Tag", "Abs Amount", "Date", "Merchant"],
        ascending=[True, False, False, True],
        kind="stable",
    )


def detail_by_merchant(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=detail_columns(df))

    return df[detail_columns(df)].sort_values(
        ["Tag", "Merchant", "Date", "Abs Amount"],
        ascending=[True, True, False, False],
        kind="stable",
    )


def detail_by_date(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=detail_columns(df))

    return df[detail_columns(df)].sort_values(
        ["Tag", "Date", "Abs Amount", "Merchant"],
        ascending=[True, False, False, True],
        kind="stable",
    )


def build_sheets(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    years = sorted(int(year) for year in df["Year"].dropna().unique())[::-1]
    tagged_df = explode_tags(df)

    return {
        "Tag Summary": append_summary_total(tag_summary(tagged_df)),
        "Tag Year Pivot": tag_year_pivot(tagged_df, years),
        "Tag Year Summary": append_summary_total(
            summary_by_dimensions(tagged_df, ["Tag", "Year"])
        ),
        "Tag Merchant Summary": append_summary_total(
            summary_by_dimensions(tagged_df, ["Tag", "Merchant"])
        ),
        "Tag Category Summary": append_summary_total(
            summary_by_dimensions(tagged_df, ["Tag", "Category"])
        ),
        "Detail By Amount": detail_by_amount(tagged_df),
        "Detail By Merchant": detail_by_merchant(tagged_df),
        "Detail By Date": detail_by_date(tagged_df),
    }


def format_workbook(writer: pd.ExcelWriter) -> None:
    for ws in writer.sheets.values():
        headers = {cell.column: str(cell.value or "").strip() for cell in ws[1]}
        first_data_column = next(
            (cell.column for cell in ws[1] if headers.get(cell.column) == "Date"),
            1,
        )
        ws.freeze_panes = f"{get_column_letter(first_data_column)}2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        for row in ws.iter_rows(min_row=2):
            is_total_row = str(row[0].value or "").strip() == TOTAL_LABEL
            for cell in row:
                cell.font = Font(name="Consolas", bold=is_total_row)
                header_name = headers.get(cell.column, "")
                if "Amount" in header_name or header_name.isdigit():
                    cell.number_format = AMOUNT_FORMAT
                elif header_name in {"Date", "First Date", "Last Date"}:
                    cell.number_format = DATE_FORMAT

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 10),
                50,
            )


def is_locked_file_error(error: OSError) -> bool:
    return isinstance(error, PermissionError) or getattr(error, "winerror", None) in {
        32,
        33,
    }


def write_workbook(output_path: Path, sheets: dict[str, pd.DataFrame]) -> bool:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, sheet_df in sheets.items():
                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
            format_workbook(writer)
    except OSError as e:
        if not is_locked_file_error(e):
            raise
        print(f"WARNING: Could not write {output_path}.")
        print("         It may be open in Excel. Close it and run the report again.")
        return False

    return True


def main() -> None:
    args = parse_args()
    transactions_df = read_csv(args.transactions)
    prepared_df = prepare_transactions(transactions_df, args.transactions)
    start_date, end_date = date_window(
        prepared_df,
        lookback_years=args.lookback_years,
        start_date=args.start_date,
        end_date=args.end_date,
    )
    filtered_df = filter_date_window(prepared_df, start_date, end_date)
    tagged_df = explode_tags(filtered_df)
    sheets = build_sheets(filtered_df)

    print(f"Read {len(prepared_df)} dated transaction rows from {args.transactions}")
    print(
        "Date window: "
        f"{start_date.strftime('%Y-%m-%d')} through {end_date.strftime('%Y-%m-%d')}"
    )
    print(f"Rows in date window: {len(filtered_df)}")
    print(f"Tagged transaction-tag rows: {len(tagged_df)}")
    print(f"Distinct tags: {tagged_df['Tag'].nunique() if not tagged_df.empty else 0}")
    print(f"Workbook tabs: {', '.join(sheets)}")
    if not write_workbook(args.output, sheets):
        return

    print(f"Saved tag activity workbook to {args.output}")


if __name__ == "__main__":
    main()
