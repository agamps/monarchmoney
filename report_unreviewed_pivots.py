import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
DEFAULT_TRANSACTIONS = Path("data/unreviewed_transactions.csv")
DEFAULT_ACCOUNT_GROUPS = Path("data/account_groups.csv")
DEFAULT_OUTPUT = Path("data/unreviewed_pivots.xlsx")

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Consolas", bold=True, color="FFFFFF")
AMOUNT_FORMAT = "#,##0.00"
DATE_FORMAT = "mm/dd/yyyy"
TOTAL_LABEL = "Total"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel pivot-style workbook from unreviewed transactions."
    )
    parser.add_argument(
        "--transactions",
        "--pending-transactions",
        dest="transactions",
        type=Path,
        default=DEFAULT_TRANSACTIONS,
        help=f"Path to pending/unreviewed transactions CSV (default: {DEFAULT_TRANSACTIONS}).",
    )
    parser.add_argument(
        "--account-groups",
        type=Path,
        default=DEFAULT_ACCOUNT_GROUPS,
        help=f"Path to account_groups.csv (default: {DEFAULT_ACCOUNT_GROUPS}).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Path for the generated Excel workbook.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> pd.DataFrame:
    last_error: UnicodeDecodeError | None = None

    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except UnicodeDecodeError as e:
            last_error = e

    assert last_error is not None
    raise ValueError(
        f"Could not decode {path} using supported encodings: {', '.join(CSV_ENCODINGS)}"
    ) from last_error


def require_columns(df: pd.DataFrame, path: Path, columns: list[str]) -> None:
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(missing)}")


def parse_amount(value: object) -> float:
    text = str(value or "").strip()
    if text == "":
        return 0.0

    is_parenthesized = text.startswith("(") and text.endswith(")")
    if is_parenthesized:
        text = text[1:-1].strip()

    text = text.replace(",", "").replace("$", "")
    amount = pd.to_numeric(text, errors="coerce")
    if pd.isna(amount):
        return 0.0

    amount = float(amount)
    return -abs(amount) if is_parenthesized else amount


def prepare_transactions(df: pd.DataFrame, path: Path) -> pd.DataFrame:
    require_columns(
        df,
        path,
        ["Transaction ID", "Account", "Merchant", "Category", "Amount"],
    )

    prepared = df.copy()
    prepared["Amount"] = prepared["Amount"].map(parse_amount)
    prepared["Total Abs Amount"] = prepared["Amount"].abs()

    if "Date" in prepared.columns:
        prepared["Date"] = pd.to_datetime(prepared["Date"], errors="coerce")

    for column in ["Account", "Merchant", "Category"]:
        prepared[column] = (
            prepared[column]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "(blank)")
        )

    return prepared


def clean_dimension(series: pd.Series, blank_label: str = "(blank)") -> pd.Series:
    return (
        series.fillna("")
        .astype(str)
        .str.strip()
        .replace("", blank_label)
    )


def prepare_account_metadata(df: pd.DataFrame, path: Path) -> pd.DataFrame:
    require_columns(df, path, ["Account", "Account Subtype"])

    prepared = df.copy()
    prepared["Account"] = clean_dimension(prepared["Account"])
    prepared["Account Subtype"] = clean_dimension(prepared["Account Subtype"])

    if "Account ID" in prepared.columns:
        prepared["Account ID"] = clean_dimension(prepared["Account ID"])

    return prepared


def add_account_subtypes(
    transactions_df: pd.DataFrame,
    account_groups_df: pd.DataFrame,
) -> pd.DataFrame:
    transactions_df = transactions_df.copy()
    join_column = (
        "Account ID"
        if "Account ID" in transactions_df.columns and "Account ID" in account_groups_df.columns
        else "Account"
    )
    transactions_df[join_column] = clean_dimension(transactions_df[join_column])

    mapping = account_groups_df[[join_column, "Account Subtype"]].drop_duplicates(
        subset=[join_column],
        keep="first",
    )

    if "Account Subtype" in transactions_df.columns:
        transactions_df = transactions_df.drop(columns=["Account Subtype"])

    merged = transactions_df.merge(mapping, on=join_column, how="left")
    merged["Account Subtype"] = clean_dimension(merged["Account Subtype"], "(unmapped)")
    return merged


def summary_by_dimensions(df: pd.DataFrame, dimensions: list[str]) -> pd.DataFrame:
    columns = dimensions + ["Transaction Count", "Net Amount", "Total Abs Amount"]
    if df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        df.groupby(dimensions, dropna=False)
        .agg(
            **{
                "Transaction Count": ("Transaction ID", "count"),
                "Net Amount": ("Amount", "sum"),
                "Total Abs Amount": ("Total Abs Amount", "sum"),
            }
        )
        .reset_index()
    )
    return summary.sort_values(
        ["Transaction Count", "Total Abs Amount"] + dimensions,
        ascending=[False, False] + [True] * len(dimensions),
        kind="stable",
    )


def summary_by(df: pd.DataFrame, dimension: str) -> pd.DataFrame:
    return summary_by_dimensions(df, [dimension])


def append_total_row(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    first_column = str(df.columns[0])
    if str(df.iloc[-1][first_column]) == TOTAL_LABEL:
        return df

    total_row = {column: "" for column in df.columns}
    total_row[first_column] = TOTAL_LABEL

    for column in df.columns:
        if column == first_column:
            continue
        if pd.api.types.is_numeric_dtype(df[column]) and not pd.api.types.is_bool_dtype(
            df[column]
        ):
            total_row[column] = df[column].sum()

    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def summary_by_pair(df: pd.DataFrame, first: str, second: str) -> pd.DataFrame:
    return summary_by_dimensions(df, [first, second])


def count_matrix(df: pd.DataFrame, index: str, columns: str) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=[index, "Total"])

    matrix = pd.pivot_table(
        df,
        index=index,
        columns=columns,
        values="Transaction ID",
        aggfunc="count",
        fill_value=0,
        margins=True,
        margins_name="Total",
    )

    if "Total" in matrix.index:
        total_row = matrix.loc[["Total"]]
        body = matrix.drop(index="Total").sort_values(
            ["Total"],
            ascending=[False],
            kind="stable",
        )
        matrix = pd.concat([body, total_row])
    else:
        matrix = matrix.sort_index(kind="stable")

    matrix = matrix.reset_index()
    matrix.columns = [str(column) for column in matrix.columns]
    return matrix


def raw_unreviewed(df: pd.DataFrame) -> pd.DataFrame:
    preferred_columns = [
        "Transaction ID",
        "Account Subtype",
        "Account",
        "Date",
        "Merchant",
        "Plaid Name",
        "Amount",
        "Category",
        "Tags",
        "Notes",
        "Hide From Reports",
        "Needs Review",
    ]
    ordered_columns = [
        column for column in preferred_columns if column in df.columns
    ] + [column for column in df.columns if column not in preferred_columns]

    sort_columns = [column for column in ["Account", "Merchant", "Date"] if column in df.columns]
    if sort_columns:
        return df[ordered_columns].sort_values(sort_columns, kind="stable")
    return df[ordered_columns]


def format_workbook(writer: pd.ExcelWriter) -> None:
    for ws in writer.sheets.values():
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = {cell.column: str(cell.value or "").strip() for cell in ws[1]}
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        for row in ws.iter_rows(min_row=2):
            is_total_row = str(row[0].value or "").strip() == TOTAL_LABEL
            for cell in row:
                cell.font = Font(name="Consolas", bold=is_total_row)
                header_name = headers.get(cell.column, "")
                if "Amount" in header_name:
                    cell.number_format = AMOUNT_FORMAT
                elif header_name == "Date":
                    cell.number_format = DATE_FORMAT

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 10),
                50,
            )


def is_locked_file_error(error: OSError) -> bool:
    return isinstance(error, PermissionError) or getattr(error, "winerror", None) in {
        32,
        33,
    }


def write_report(output_path: Path, df: pd.DataFrame) -> bool:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sheets = {
        "Merchant Summary": append_total_row(summary_by(df, "Merchant")),
        "Account Summary": append_total_row(summary_by(df, "Account")),
        "Category Summary": append_total_row(summary_by(df, "Category")),
        "Acct Subtype Account Merchant": append_total_row(
            summary_by_dimensions(df, ["Account Subtype", "Account", "Merchant"])
        ),
        "Merchant Account": append_total_row(summary_by_pair(df, "Merchant", "Account")),
        "Account Merchant": append_total_row(summary_by_pair(df, "Account", "Merchant")),
        "Account Category": append_total_row(summary_by_pair(df, "Account", "Category")),
        "Merchant Acct Pivot": count_matrix(df, "Merchant", "Account"),
        "Account Merch Pivot": count_matrix(df, "Account", "Merchant"),
        "Account Cat Pivot": count_matrix(df, "Account", "Category"),
        "Raw Unreviewed": append_total_row(raw_unreviewed(df)),
    }

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, sheet_df in sheets.items():
                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
            format_workbook(writer)
    except OSError as e:
        if not is_locked_file_error(e):
            raise
        print(f"WARNING: Could not write {output_path}.")
        print("         It may be open in Excel. Close it and run the report again.")
        return False

    return True


def main() -> None:
    args = parse_args()
    transactions_df = read_csv(args.transactions)
    account_groups_df = read_csv(args.account_groups)
    prepared_df = prepare_transactions(transactions_df, args.transactions)
    prepared_account_groups_df = prepare_account_metadata(
        account_groups_df,
        args.account_groups,
    )
    prepared_df = add_account_subtypes(prepared_df, prepared_account_groups_df)

    print(f"Read {len(prepared_df)} unreviewed transaction rows from {args.transactions}")
    print(f"Read account subtype mappings from {args.account_groups}")
    if not write_report(args.output, prepared_df):
        return

    print(f"Saved unreviewed pivot workbook to {args.output}")


if __name__ == "__main__":
    main()
