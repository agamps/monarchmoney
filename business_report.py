import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
DEFAULT_TRANSACTIONS = Path("data/all_transactions.csv")
DEFAULT_GROUPS = Path("data/category_groups.csv")

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Consolas", bold=True, color="FFFFFF")
AMOUNT_FORMAT = "#,##0.00"
DATE_FORMAT = "mm/dd/yyyy"
TOTAL_LABEL = "Total"
BLANK_LABEL = "(blank)"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel report for selected category groups and year."
    )
    parser.add_argument(
        "--group-terms",
        help=(
            "Comma-separated group name search terms, such as 'Virtus, Immanent'. "
            "Prefix a term with '-' or '!' to exclude it, such as 'Immanent, -Income'. "
            "A term matches any category group containing that text."
        ),
    )
    parser.add_argument(
        "--exclude-group-terms",
        help=(
            "Comma-separated group name search terms to exclude after applying "
            "--group-terms."
        ),
    )
    parser.add_argument(
        "--fiscal-year",
        type=int,
        help="Fiscal year to report on, using the transaction date year.",
    )
    parser.add_argument(
        "--transactions",
        type=Path,
        default=DEFAULT_TRANSACTIONS,
        help="Path to the transactions CSV.",
    )
    parser.add_argument(
        "--groups",
        type=Path,
        default=DEFAULT_GROUPS,
        help="Path to the category groups CSV.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output Excel file path. Defaults to data/<group_terms>_business_report_<year>.xlsx.",
    )
    return parser.parse_args()


def split_terms(value: str | None) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in re.split(r"[,;\n]", value) if part.strip()]


def dedupe_terms(terms: list[str]) -> list[str]:
    deduped: list[str] = []
    seen: set[str] = set()
    for term in terms:
        key = term.casefold()
        if key in seen:
            continue
        deduped.append(term)
        seen.add(key)
    return deduped


def prompt_for_group_filters(
    include_value: str | None,
    exclude_value: str | None,
) -> tuple[list[str], list[str]]:
    raw = include_value
    while raw is None or raw.strip() == "":
        raw = input("Group terms, comma separated: ").strip()

    include_terms: list[str] = []
    exclude_terms: list[str] = []
    for term in split_terms(raw):
        if term.startswith(("-", "!")):
            excluded_term = term[1:].strip()
            if excluded_term:
                exclude_terms.append(excluded_term)
        else:
            include_terms.append(term)

    for term in split_terms(exclude_value):
        excluded_term = term[1:].strip() if term.startswith(("-", "!")) else term
        if excluded_term:
            exclude_terms.append(excluded_term)

    include_terms = dedupe_terms(include_terms)
    exclude_terms = dedupe_terms(exclude_terms)
    if not include_terms:
        raise ValueError("Enter at least one group include term.")

    return include_terms, exclude_terms


def prompt_for_year(value: int | None) -> int:
    if value is not None:
        return value

    while True:
        raw = input("Fiscal year: ").strip()
        if raw.isdigit() and len(raw) == 4:
            return int(raw)
        print("Enter a 4-digit year like 2026.")


def slugify(parts: list[str]) -> str:
    text = "_".join(parts).casefold()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "group_subset"


def default_output_path(
    include_terms: list[str],
    exclude_terms: list[str],
    fiscal_year: int,
) -> Path:
    slug_parts = [*include_terms, *(f"exclude_{term}" for term in exclude_terms)]
    return Path("data") / f"{slugify(slug_parts)}_business_report_{fiscal_year}.xlsx"


def read_csv(path: Path) -> pd.DataFrame:
    last_error: UnicodeDecodeError | None = None

    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except UnicodeDecodeError as e:
            last_error = e

    assert last_error is not None
    raise ValueError(
        f"Could not decode {path} using supported encodings: {', '.join(CSV_ENCODINGS)}"
    ) from last_error


def require_columns(df: pd.DataFrame, path: Path, columns: list[str]) -> None:
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(missing)}")


def parse_amount(value: object) -> float:
    text = str(value or "").strip()
    if text == "":
        return 0.0

    is_parenthesized = text.startswith("(") and text.endswith(")")
    if is_parenthesized:
        text = text[1:-1].strip()

    text = text.replace(",", "").replace("$", "")
    amount = pd.to_numeric(text, errors="coerce")
    if pd.isna(amount):
        return float("nan")

    amount = float(amount)
    return -abs(amount) if is_parenthesized else amount


def parse_dates(values: pd.Series) -> pd.Series:
    try:
        return pd.to_datetime(values, format="mixed", errors="coerce")
    except TypeError:
        return pd.to_datetime(values, errors="coerce")


def clean_dimension(values: pd.Series) -> pd.Series:
    return (
        values.fillna("")
        .astype(str)
        .str.strip()
        .replace("", BLANK_LABEL)
    )


def select_group_names(
    groups_df: pd.DataFrame,
    include_terms: list[str],
    exclude_terms: list[str],
) -> list[str]:
    unique_groups = (
        groups_df["Group Name"]
        .dropna()
        .astype(str)
        .str.strip()
        .loc[lambda values: values != ""]
        .drop_duplicates()
        .tolist()
    )

    selected: list[str] = []
    selected_keys: set[str] = set()
    for term in include_terms:
        normalized_term = term.casefold()
        for group_name in unique_groups:
            if normalized_term not in group_name.casefold():
                continue
            key = group_name.casefold()
            if key in selected_keys:
                continue
            selected.append(group_name)
            selected_keys.add(key)

    if exclude_terms:
        selected = [
            group_name
            for group_name in selected
            if not any(
                exclude_term.casefold() in group_name.casefold()
                for exclude_term in exclude_terms
            )
        ]

    if not selected:
        available = ", ".join(unique_groups)
        exclude_message = (
            f" after excluding: {', '.join(exclude_terms)}"
            if exclude_terms
            else ""
        )
        raise ValueError(
            f"No category groups matched: {', '.join(include_terms)}{exclude_message}. "
            f"Available groups: {available}"
        )

    return sorted(selected, key=str.casefold)


def prepare_transactions(
    transactions_df: pd.DataFrame,
    groups_df: pd.DataFrame,
    transactions_path: Path,
    groups_path: Path,
    fiscal_year: int,
    selected_groups: list[str],
) -> pd.DataFrame:
    require_columns(transactions_df, transactions_path, ["Date", "Amount", "Category"])
    require_columns(groups_df, groups_path, ["Category Name", "Group Name"])

    tx = transactions_df.copy()
    tx["Date"] = parse_dates(tx["Date"])
    tx["Amount"] = tx["Amount"].map(parse_amount)
    tx = tx.dropna(subset=["Date", "Amount", "Category"])
    tx = tx[tx["Date"].dt.year == fiscal_year]

    for column in ["Account", "Merchant", "Category"]:
        if column not in tx.columns:
            tx[column] = BLANK_LABEL
        tx[column] = clean_dimension(tx[column])

    group_lookup = groups_df[["Category Name", "Group Name"]].dropna().copy()
    group_lookup["Category Name"] = group_lookup["Category Name"].astype(str).str.strip()
    group_lookup["Group Name"] = group_lookup["Group Name"].astype(str).str.strip()

    merged = tx.merge(
        group_lookup,
        how="left",
        left_on="Category",
        right_on="Category Name",
    )
    filtered = merged[merged["Group Name"].isin(selected_groups)].copy()
    filtered["Group Name"] = pd.Categorical(
        filtered["Group Name"],
        categories=selected_groups,
        ordered=True,
    )
    filtered["Total Abs Amount"] = filtered["Amount"].abs()
    return filtered


def append_total_row(df: pd.DataFrame, first_column: str) -> pd.DataFrame:
    if df.empty:
        return df

    total_row = {column: "" for column in df.columns}
    total_row[first_column] = TOTAL_LABEL
    for column in df.columns:
        if column == first_column:
            continue
        if pd.api.types.is_numeric_dtype(df[column]) and not pd.api.types.is_bool_dtype(
            df[column]
        ):
            total_row[column] = df[column].sum()

    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def summary_by(df: pd.DataFrame, dimensions: list[str]) -> pd.DataFrame:
    columns = dimensions + ["Transaction Count", "Net Amount", "Total Abs Amount"]
    if df.empty:
        return pd.DataFrame(columns=columns)

    summary = (
        df.groupby(dimensions, observed=True)
        .agg(
            **{
                "Transaction Count": ("Amount", "count"),
                "Net Amount": ("Amount", "sum"),
                "Total Abs Amount": ("Total Abs Amount", "sum"),
            }
        )
        .reset_index()
    )
    summary = summary.sort_values(
        dimensions,
        key=lambda column: column.astype(str).str.casefold(),
        kind="stable",
    )
    return append_total_row(summary, dimensions[0])


def group_summary(df: pd.DataFrame, selected_groups: list[str]) -> pd.DataFrame:
    columns = [
        "Group Name",
        "Transaction Count",
        "Net Amount",
        "Total Abs Amount",
    ]
    if df.empty:
        summary = pd.DataFrame(columns=columns)
    else:
        summary = (
            df.groupby("Group Name", observed=True)
            .agg(
                **{
                    "Transaction Count": ("Amount", "count"),
                    "Net Amount": ("Amount", "sum"),
                    "Total Abs Amount": ("Total Abs Amount", "sum"),
                }
            )
            .reset_index()
        )

    zero_rows = pd.DataFrame(
        {
            "Group Name": selected_groups,
            "Transaction Count": [0] * len(selected_groups),
            "Net Amount": [0.0] * len(selected_groups),
            "Total Abs Amount": [0.0] * len(selected_groups),
        }
    )
    summary = (
        zero_rows.set_index("Group Name")
        .combine_first(summary.set_index("Group Name"))
        .loc[selected_groups]
        .reset_index()
    )
    return append_total_row(summary, "Group Name")


def tagged_transactions(df: pd.DataFrame) -> pd.DataFrame:
    if "Tags" not in df.columns:
        return pd.DataFrame(columns=[*df.columns, "Tag"])

    tagged = df.copy()
    tagged["Tag"] = tagged["Tags"].fillna("").astype(str).str.split(",")
    tagged = tagged.explode("Tag")
    tagged["Tag"] = tagged["Tag"].fillna("").astype(str).str.strip()
    tagged = tagged[tagged["Tag"] != ""]
    return tagged


def transactions_detail(df: pd.DataFrame) -> pd.DataFrame:
    transactions_tab = df.copy()
    transactions_tab = transactions_tab.drop(
        columns=["Category Name", "Total Abs Amount", "Hide From Reports", "Needs Review"],
        errors="ignore",
    )
    transactions_tab = transactions_tab.rename(columns={"Category": "Category Name"})
    transactions_tab = transactions_tab.sort_values(
        ["Group Name", "Category Name", "Date", "Merchant"],
        kind="stable",
    )

    first_columns = [
        "Transaction ID",
        "Group Name",
        "Category Name",
        "Account",
        "Date",
        "Merchant",
        "Amount",
        "Tags",
        "Notes",
    ]
    ordered_columns = [
        column for column in first_columns if column in transactions_tab.columns
    ] + [column for column in transactions_tab.columns if column not in first_columns]

    return transactions_tab[ordered_columns]


def build_sheets(df: pd.DataFrame, selected_groups: list[str]) -> dict[str, pd.DataFrame]:
    tagged_df = tagged_transactions(df)
    sheets = {
        "Group": group_summary(df, selected_groups),
        "Group Category": summary_by(df, ["Group Name", "Category"]),
        "Merchant": summary_by(df, ["Merchant"]),
        "Category": summary_by(df, ["Category"]),
        "Account": summary_by(df, ["Account"]),
        "Account Group": summary_by(df, ["Account", "Group Name"]),
        "Account Category": summary_by(df, ["Account", "Category"]),
        "Merchant Category": summary_by(df, ["Merchant", "Category"]),
        "Merchant Account": summary_by(df, ["Merchant", "Account"]),
    }

    if not tagged_df.empty:
        sheets["Tag"] = summary_by(tagged_df, ["Tag"])
        sheets["Tag Category"] = summary_by(tagged_df, ["Tag", "Category"])

    sheets["Transactions"] = transactions_detail(df)
    return sheets


def format_workbook(writer: pd.ExcelWriter) -> None:
    for ws in writer.sheets.values():
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = {cell.column: str(cell.value or "").strip() for cell in ws[1]}
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        for row in ws.iter_rows(min_row=2):
            is_total_row = str(row[0].value or "").strip() == TOTAL_LABEL
            for cell in row:
                cell.font = Font(name="Consolas", bold=is_total_row)
                header_name = headers.get(cell.column, "")
                if header_name in {"Amount", "Net Amount", "Total Abs Amount"}:
                    cell.number_format = AMOUNT_FORMAT
                elif header_name == "Date":
                    cell.number_format = DATE_FORMAT

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
                max(max_length + 2, 10),
                50,
            )


def is_locked_file_error(error: OSError) -> bool:
    return isinstance(error, PermissionError) or getattr(error, "winerror", None) in {
        32,
        33,
    }


def write_excel_report(output_path: Path, sheets: dict[str, pd.DataFrame]) -> bool:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, sheet_df in sheets.items():
                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
            format_workbook(writer)
    except OSError as e:
        if not is_locked_file_error(e):
            raise
        print(f"WARNING: Could not write {output_path}.")
        print("         It may be open in Excel. Close it and run the report again.")
        return False

    return True


def main() -> None:
    args = parse_args()
    include_terms, exclude_terms = prompt_for_group_filters(
        args.group_terms,
        args.exclude_group_terms,
    )
    fiscal_year = prompt_for_year(args.fiscal_year)
    output_path = args.output or default_output_path(
        include_terms,
        exclude_terms,
        fiscal_year,
    )

    transactions_df = read_csv(args.transactions)
    groups_df = read_csv(args.groups)
    require_columns(groups_df, args.groups, ["Category Name", "Group Name"])
    selected_groups = select_group_names(groups_df, include_terms, exclude_terms)
    filtered_df = prepare_transactions(
        transactions_df,
        groups_df,
        args.transactions,
        args.groups,
        fiscal_year,
        selected_groups,
    )
    sheets = build_sheets(filtered_df, selected_groups)

    print(f"Fiscal year: {fiscal_year}")
    print(f"Group include terms: {', '.join(include_terms)}")
    if exclude_terms:
        print(f"Group exclude terms: {', '.join(exclude_terms)}")
    print(f"Groups included: {', '.join(selected_groups)}")
    print(f"Transactions included: {len(filtered_df)}")
    if not write_excel_report(output_path, sheets):
        return

    print(f"Saved Excel report to {output_path}")


if __name__ == "__main__":
    main()
