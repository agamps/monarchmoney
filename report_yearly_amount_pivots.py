import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
DEFAULT_TRANSACTIONS = Path("data/all_transactions.csv")
DEFAULT_GROUPS = Path("data/category_groups.csv")
DEFAULT_OUTPUT = Path("data/yearly_amount_pivots.xlsx")

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Consolas", bold=True, color="FFFFFF")
AMOUNT_FORMAT = "#,##0.00"
TOTAL_LABEL = "Total"
BLANK_LABEL = "(blank)"
UNMAPPED_GROUP = "Unmapped"


def optional_years_input(value: str | None) -> str | None:
    if value is None:
        return None

    text = value.strip()
    if text == "" or text.casefold() in {"auto", "default", "none"}:
        return None
    return text


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel workbook of amount pivots by year."
    )
    parser.add_argument(
        "--transactions",
        type=Path,
        default=DEFAULT_TRANSACTIONS,
        help="Source transactions CSV. Defaults to data/all_transactions.csv.",
    )
    parser.add_argument(
        "--groups",
        type=Path,
        default=DEFAULT_GROUPS,
        help="Category groups CSV. Defaults to data/category_groups.csv.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Path for the generated Excel workbook.",
    )
    parser.add_argument(
        "--years",
        type=optional_years_input,
        default=None,
        help="Optional comma-separated list of year columns to include, in order.",
    )
    parser.add_argument(
        "--exclude-groups",
        nargs="?",
        const="",
        default="",
        help="Optional comma-separated group name terms to exclude.",
    )
    parser.add_argument(
        "--exclude-categories",
        nargs="?",
        const="",
        default="",
        help="Optional comma-separated category name terms to exclude.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> pd.DataFrame:
    last_error: UnicodeDecodeError | None = None

    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except UnicodeDecodeError as e:
            last_error = e

    assert last_error is not None
    raise ValueError(
        f"Could not decode {path} using supported encodings: {', '.join(CSV_ENCODINGS)}"
    ) from last_error


def require_columns(df: pd.DataFrame, path: Path, columns: list[str]) -> None:
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(missing)}")


def parse_amount(value: object) -> float:
    text = str(value or "").strip()
    if text == "":
        return 0.0

    is_parenthesized = text.startswith("(") and text.endswith(")")
    if is_parenthesized:
        text = text[1:-1].strip()

    text = text.replace(",", "").replace("$", "")
    amount = pd.to_numeric(text, errors="coerce")
    if pd.isna(amount):
        return 0.0

    amount = float(amount)
    return -abs(amount) if is_parenthesized else amount


def parse_dates(values: pd.Series) -> pd.Series:
    try:
        return pd.to_datetime(values, format="mixed", errors="coerce")
    except TypeError:
        return pd.to_datetime(values, errors="coerce")


def clean_dimension(values: pd.Series) -> pd.Series:
    return (
        values.fillna("")
        .astype(str)
        .str.strip()
        .replace("", BLANK_LABEL)
    )


def load_category_group_map(path: Path) -> dict[str, str]:
    if not path.exists():
        print(f"WARNING: {path} not found. Group tabs will use {UNMAPPED_GROUP!r}.")
        return {}

    groups_df = read_csv(path)
    require_columns(groups_df, path, ["Category Name", "Group Name"])

    groups_df = groups_df[["Category Name", "Group Name"]].dropna().copy()
    groups_df["Category Name"] = groups_df["Category Name"].astype(str).str.strip()
    groups_df["Group Name"] = groups_df["Group Name"].astype(str).str.strip()
    groups_df = groups_df[
        (groups_df["Category Name"] != "") & (groups_df["Group Name"] != "")
    ]

    return dict(zip(groups_df["Category Name"], groups_df["Group Name"], strict=False))


def prepare_transactions(
    transactions_df: pd.DataFrame,
    transactions_path: Path,
    category_to_group: dict[str, str],
) -> pd.DataFrame:
    require_columns(transactions_df, transactions_path, ["Date", "Amount", "Category"])

    prepared = transactions_df.copy()
    prepared["Date"] = parse_dates(prepared["Date"])
    prepared["Amount"] = prepared["Amount"].map(parse_amount)
    prepared = prepared.dropna(subset=["Date"])
    prepared["Year"] = prepared["Date"].dt.year.astype(int)

    for column in ["Account", "Merchant", "Category"]:
        if column not in prepared.columns:
            prepared[column] = BLANK_LABEL
        prepared[column] = clean_dimension(prepared[column])

    prepared["Group"] = (
        prepared["Category"]
        .map(category_to_group)
        .fillna(UNMAPPED_GROUP)
        .astype(str)
        .str.strip()
        .replace("", UNMAPPED_GROUP)
    )

    return prepared


def year_columns(df: pd.DataFrame) -> list[int]:
    return sorted(int(year) for year in df["Year"].dropna().unique())[::-1]


def split_terms(value: str | None) -> list[str]:
    if value is None:
        return []

    terms = [part.strip() for part in re.split(r"[,;\n]", value) if part.strip()]
    return [term for term in terms if term.casefold() not in {"none", "default"}]


def term_mask(values: pd.Series, terms: list[str]) -> pd.Series:
    if not terms:
        return pd.Series(False, index=values.index)

    normalized = values.fillna("").astype(str).str.casefold()
    mask = pd.Series(False, index=values.index)
    for term in terms:
        mask = mask | normalized.str.contains(re.escape(term.casefold()), na=False)
    return mask


def apply_exclusions(
    df: pd.DataFrame,
    exclude_groups: list[str],
    exclude_categories: list[str],
) -> tuple[pd.DataFrame, int]:
    excluded_mask = term_mask(df["Group"], exclude_groups) | term_mask(
        df["Category"],
        exclude_categories,
    )
    return df.loc[~excluded_mask].copy(), int(excluded_mask.sum())


def parse_year_names(values: list[object], source: str) -> list[int]:
    years: list[int] = []
    seen: set[int] = set()
    for value in values:
        text = str(value or "").strip()
        if text == "" or text.startswith("#"):
            continue

        for match in re.findall(r"\d{2,4}", text):
            year = int(match)
            if year < 100:
                year += 2000
            if year not in seen:
                years.append(year)
                seen.add(year)

    if not years:
        raise ValueError(f"{source} does not contain any year names.")

    return years


def parse_year_columns_from_input(
    years_input: str | None,
) -> tuple[list[int] | None, str | None]:
    if years_input is None:
        return None, None

    return parse_year_names([years_input], "year list input"), "year list input"


def amount_pivot(
    df: pd.DataFrame,
    dimensions: list[str],
    years: list[int],
) -> pd.DataFrame:
    year_labels = [str(year) for year in years]
    columns = dimensions + year_labels + [TOTAL_LABEL]
    if df.empty:
        return pd.DataFrame(columns=columns)

    pivot = pd.pivot_table(
        df,
        index=dimensions,
        columns="Year",
        values="Amount",
        aggfunc="sum",
        fill_value=0.0,
    )

    for year in years:
        if year not in pivot.columns:
            pivot[year] = 0.0

    pivot = pivot[years]
    pivot[TOTAL_LABEL] = pivot.sum(axis=1)
    pivot = pivot.reset_index()
    pivot.columns = [str(column) for column in pivot.columns]

    pivot = pivot.sort_values(
        dimensions,
        key=lambda column: column.astype(str).str.casefold(),
        kind="stable",
    )

    total_row = {column: "" for column in pivot.columns}
    total_row[dimensions[0]] = TOTAL_LABEL
    for column in year_labels + [TOTAL_LABEL]:
        total_row[column] = pivot[column].sum()

    return pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)


def tagged_transactions(df: pd.DataFrame) -> pd.DataFrame:
    if "Tags" not in df.columns:
        return pd.DataFrame(columns=[*df.columns, "Tag"])

    tagged = df.copy()
    tagged["Tag"] = tagged["Tags"].fillna("").astype(str).str.split(",")
    tagged = tagged.explode("Tag")
    tagged["Tag"] = tagged["Tag"].fillna("").astype(str).str.strip()
    tagged = tagged[tagged["Tag"] != ""]
    return tagged


def build_sheets(df: pd.DataFrame, years: list[int]) -> dict[str, pd.DataFrame]:
    tagged_df = tagged_transactions(df)

    sheets = {
        "Group": amount_pivot(df, ["Group"], years),
        "Group Category": amount_pivot(df, ["Group", "Category"], years),
        "Merchant": amount_pivot(df, ["Merchant"], years),
        "Category": amount_pivot(df, ["Category"], years),
        "Account": amount_pivot(df, ["Account"], years),
        "Account Group": amount_pivot(df, ["Account", "Group"], years),
        "Account Category": amount_pivot(df, ["Account", "Category"], years),
        "Merchant Category": amount_pivot(df, ["Merchant", "Category"], years),
        "Merchant Account": amount_pivot(df, ["Merchant", "Account"], years),
    }

    if not tagged_df.empty:
        sheets["Tag"] = amount_pivot(tagged_df, ["Tag"], years)
        sheets["Tag Category"] = amount_pivot(tagged_df, ["Tag", "Category"], years)

    return sheets


def format_workbook(writer: pd.ExcelWriter) -> None:
    for ws in writer.sheets.values():
        headers = {cell.column: str(cell.value or "").strip() for cell in ws[1]}
        first_amount_column = next(
            (
                cell.column
                for cell in ws[1]
                if headers.get(cell.column, "").isdigit()
                or headers.get(cell.column, "") == TOTAL_LABEL
            ),
            2,
        )
        ws.freeze_panes = f"{get_column_letter(first_amount_column)}2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL

        for row in ws.iter_rows(min_row=2):
            is_total_row = str(row[0].value or "").strip() == TOTAL_LABEL
            for cell in row:
                cell.font = Font(name="Consolas", bold=is_total_row)
                header_name = headers.get(cell.column, "")
                if header_name.isdigit() or header_name == TOTAL_LABEL:
                    cell.number_format = AMOUNT_FORMAT

        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 10),
                50,
            )


def is_locked_file_error(error: OSError) -> bool:
    return isinstance(error, PermissionError) or getattr(error, "winerror", None) in {
        32,
        33,
    }


def write_workbook(output_path: Path, sheets: dict[str, pd.DataFrame]) -> bool:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, sheet_df in sheets.items():
                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
            format_workbook(writer)
    except OSError as e:
        if not is_locked_file_error(e):
            raise
        print(f"WARNING: Could not write {output_path}.")
        print("         It may be open in Excel. Close it and run the report again.")
        return False

    return True


def main() -> None:
    args = parse_args()
    category_to_group = load_category_group_map(args.groups)
    transactions_df = read_csv(args.transactions)
    prepared_df = prepare_transactions(
        transactions_df,
        args.transactions,
        category_to_group,
    )
    prepared_count = len(prepared_df)
    exclude_groups = split_terms(args.exclude_groups)
    exclude_categories = split_terms(args.exclude_categories)
    prepared_df, excluded_count = apply_exclusions(
        prepared_df,
        exclude_groups,
        exclude_categories,
    )
    configured_years, years_source = parse_year_columns_from_input(args.years)
    years = configured_years or year_columns(prepared_df)
    sheets = build_sheets(prepared_df, years)

    print(f"Read {prepared_count} transaction rows from {args.transactions}")
    print(f"Loaded {len(category_to_group)} category-to-group mappings from {args.groups}")
    if exclude_groups:
        print(f"Excluded group terms: {', '.join(exclude_groups)}")
    if exclude_categories:
        print(f"Excluded category terms: {', '.join(exclude_categories)}")
    if exclude_groups or exclude_categories:
        print(f"Excluded transaction rows: {excluded_count}")
        print(f"Rows included after exclusions: {len(prepared_df)}")
    if years_source:
        print(f"Loaded year columns from {years_source}")
    print(f"Year columns: {', '.join(str(year) for year in years)}")
    print(f"Workbook tabs: {', '.join(sheets)}")
    if not write_workbook(args.output, sheets):
        return

    print(f"Saved yearly amount pivot workbook to {args.output}")


if __name__ == "__main__":
    main()
