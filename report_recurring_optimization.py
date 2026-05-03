import argparse
import fnmatch
import math
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
DEFAULT_TRANSACTIONS = Path("data/all_transactions.csv")
DEFAULT_GROUPS = Path("data/category_groups.csv")
DEFAULT_OPTIMIZABLE_GROUPS = Path("data/optimizable_groups.txt")
DEFAULT_OPTIMIZABLE_CATEGORIES = Path("data/optimizable_categories.txt")
DEFAULT_OUTPUT = Path("data/recurring_optimization.xlsx")

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Consolas", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Consolas")
AMOUNT_FORMAT = "#,##0.00"
DATE_FORMAT = "mm/dd/yyyy"
INTEGER_FORMAT = "0"
PERCENT_FORMAT = "0.0%"
SCORE_FORMAT = "0.0"
BLANK_LABEL = "(blank)"
UNMAPPED_GROUP = "Unmapped"
TOTAL_LABEL = "Total"

ACTION_COLUMNS = [
    "Priority",
    "Optimization Type",
    "Selection Type",
    "Selection Match",
    "Merchant",
    "Group",
    "Top Category",
    "Cadence",
    "Active Status",
    "Confidence",
    "Opportunity Score",
    "Estimated Annual Spend",
    "Trailing 12 Month Spend",
    "Recent Monthly Average",
    "Typical Amount",
    "Last Amount",
    "Last Transaction Date",
    "Potential Annual Savings 10%",
    "Potential Annual Savings 25%",
    "Potential Annual Savings 50%",
    "Potential Annual Savings 100%",
    "Price Change Since First Seen",
    "Amount Stability",
    "Recommendation",
]

SUBSCRIPTION_TERMS = {
    "subscription",
    "software",
    "saas",
    "cloud",
    "hosting",
    "storage",
    "streaming",
    "membership",
    "memberships",
    "gym",
    "fitness",
    "phone",
    "wireless",
    "internet",
    "utilities",
}
SHOPPING_TERMS = {
    "shopping",
    "retail",
    "amazon",
    "household",
    "restaurants",
    "dining",
    "coffee",
    "delivery",
    "groceries",
    "grocery",
    "gas",
    "fuel",
    "rideshare",
    "transport",
    "travel",
}
NEGOTIATION_TERMS = {
    "insurance",
    "utility",
    "utilities",
    "internet",
    "phone",
    "wireless",
    "bank",
    "fee",
    "fees",
    "interest",
    "loan",
    "rent",
    "mortgage",
}
LOW_CONTROL_TERMS = {
    "tax",
    "taxes",
    "tuition",
    "medical",
    "health",
    "healthcare",
    "donation",
    "charity",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create an expense-only Excel report of recurring merchants in "
            "user-selected optimizable category groups or category terms."
        )
    )
    parser.add_argument(
        "--transactions",
        type=Path,
        default=DEFAULT_TRANSACTIONS,
        help="Source transactions CSV. Defaults to data/all_transactions.csv.",
    )
    parser.add_argument(
        "--groups",
        type=Path,
        default=DEFAULT_GROUPS,
        help="Category groups CSV. Defaults to data/category_groups.csv.",
    )
    parser.add_argument(
        "--optimizable-type",
        choices=("groups", "categories"),
        default="groups",
        help=(
            "Use optimizable group names or wildcard category terms for the "
            "expense selection. Defaults to groups."
        ),
    )
    parser.add_argument(
        "--optimizable-groups",
        type=Path,
        default=DEFAULT_OPTIMIZABLE_GROUPS,
        help=(
            "Flat text file with one category group name per line. Only expense "
            "transactions in these groups are analyzed. Blank lines and lines "
            "starting with # are ignored. Defaults to data/optimizable_groups.txt."
        ),
    )
    parser.add_argument(
        "--optimizable-categories",
        type=Path,
        default=DEFAULT_OPTIMIZABLE_CATEGORIES,
        help=(
            "Flat text file with one category search term per line. Used when "
            "--optimizable-type categories is selected. Terms are case-insensitive "
            "wildcard/substring matches against category names. Defaults to "
            "data/optimizable_categories.txt."
        ),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Path for the generated Excel workbook.",
    )
    parser.add_argument(
        "--lookback-months",
        type=int,
        default=36,
        help="Months of history to inspect. Values above 36 are capped at 36.",
    )
    parser.add_argument(
        "--recent-months",
        type=int,
        default=6,
        help="Months used for recency-weighted averages and scoring.",
    )
    parser.add_argument(
        "--min-occurrences",
        type=int,
        default=3,
        help=(
            "Minimum transaction count for a merchant/selection to be considered "
            "recurring."
        ),
    )
    parser.add_argument(
        "--min-months",
        type=int,
        default=3,
        help=(
            "Minimum distinct months for a merchant/selection to be considered "
            "recurring."
        ),
    )
    parser.add_argument(
        "--min-confidence",
        type=float,
        default=45.0,
        help="Minimum recurrence confidence score to include in the report.",
    )
    parser.add_argument(
        "--min-annualized-spend",
        type=float,
        default=50.0,
        help="Minimum estimated annual spend to include.",
    )
    parser.add_argument(
        "--amount-tolerance",
        type=float,
        default=0.20,
        help="Typical amount variation tolerated before a merchant looks variable.",
    )
    parser.add_argument(
        "--expense-sign",
        choices=("positive", "negative"),
        default="negative",
        help=(
            "Which signed Amount values represent expenses. Monarch API exports "
            "in this workflow normally use negative expenses and positive income."
        ),
    )
    parser.add_argument(
        "--include-hidden",
        action="store_true",
        help="Include rows marked Hide From Reports=true.",
    )
    parser.add_argument(
        "--top",
        type=int,
        default=250,
        help="Maximum merchant/selection candidates to keep after scoring.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> pd.DataFrame:
    last_error: UnicodeDecodeError | None = None

    for encoding in CSV_ENCODINGS:
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except UnicodeDecodeError as e:
            last_error = e

    assert last_error is not None
    raise ValueError(
        f"Could not decode {path} using supported encodings: {', '.join(CSV_ENCODINGS)}"
    ) from last_error


def read_text_file(path: Path) -> str:
    last_error: UnicodeDecodeError | None = None

    for encoding in CSV_ENCODINGS:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError as e:
            last_error = e

    assert last_error is not None
    raise ValueError(
        f"Could not decode {path} using supported encodings: {', '.join(CSV_ENCODINGS)}"
    ) from last_error


def require_columns(df: pd.DataFrame, path: Path, columns: list[str]) -> None:
    missing = [column for column in columns if column not in df.columns]
    if missing:
        raise ValueError(f"{path} is missing required columns: {', '.join(missing)}")


def parse_amount(value: object) -> float:
    text = str(value or "").strip()
    if text == "":
        return float("nan")

    is_parenthesized = text.startswith("(") and text.endswith(")")
    if is_parenthesized:
        text = text[1:-1].strip()

    text = text.replace(",", "").replace("$", "")
    amount = pd.to_numeric(text, errors="coerce")
    if pd.isna(amount):
        return float("nan")

    amount = float(amount)
    return -abs(amount) if is_parenthesized else amount


def parse_dates(values: pd.Series) -> pd.Series:
    try:
        parsed = pd.to_datetime(values, format="mixed", errors="coerce")
        if parsed.notna().any() or values.dropna().empty:
            return parsed
    except TypeError:
        pass
    return pd.to_datetime(values, errors="coerce")


def normalize_bool(value: object) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value

    text = str(value).strip().casefold()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return None


def clean_dimension(values: pd.Series) -> pd.Series:
    return (
        values.fillna("")
        .astype(str)
        .str.strip()
        .replace("", BLANK_LABEL)
    )


def top_values(values: pd.Series, limit: int = 3) -> str:
    cleaned = (
        values.fillna("")
        .astype(str)
        .str.strip()
        .loc[lambda series: (series != "") & (series != BLANK_LABEL)]
    )
    if cleaned.empty:
        return BLANK_LABEL
    return ", ".join(cleaned.value_counts().head(limit).index.tolist())


def safe_div(numerator: float, denominator: float) -> float:
    if denominator == 0 or math.isnan(denominator):
        return 0.0
    return numerator / denominator


def clamp(value: float, low: float = 0.0, high: float = 1.0) -> float:
    return max(low, min(high, value))


def month_count(start: pd.Timestamp, end: pd.Timestamp) -> int:
    return max(1, (end.year - start.year) * 12 + (end.month - start.month) + 1)


def normalize_name(value: object) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def load_optimizable_terms(path: Path, label: str) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(
            f"{path} does not exist. Create a flat text file with one optimizable "
            f"{label} per line, then pass it to the matching optimizable argument."
        )

    terms: list[str] = []
    for line in read_text_file(path).splitlines():
        text = line.strip()
        if text == "" or text.startswith("#"):
            continue
        terms.append(text)

    if not terms:
        raise ValueError(
            f"{path} does not contain any {label}s. Add one {label} per line."
        )

    return terms


def resolve_requested_groups(
    available_groups: pd.Series,
    requested_groups: list[str],
    source_path: Path,
) -> tuple[list[str], list[str]]:
    available_by_key: dict[str, str] = {}
    for value in sorted(available_groups.dropna().astype(str).unique()):
        key = normalize_name(value)
        if key:
            available_by_key.setdefault(key, value)

    matched: list[str] = []
    missing: list[str] = []
    seen: set[str] = set()
    for group in requested_groups:
        key = normalize_name(group)
        if key in available_by_key:
            value = available_by_key[key]
            if value not in seen:
                matched.append(value)
                seen.add(value)
        else:
            missing.append(group)

    if matched:
        return matched, missing

    available_preview = ", ".join(sorted(available_by_key.values())[:25])
    raise ValueError(
        f"No groups from {source_path} matched expense transactions. "
        f"Available expense groups include: {available_preview}"
    )


def category_matches_term(category: str, term: str) -> bool:
    normalized_category = normalize_name(category)
    normalized_term = normalize_name(term)
    if normalized_term == "":
        return False
    if "*" in normalized_term or "?" in normalized_term:
        return fnmatch.fnmatchcase(normalized_category, normalized_term)
    return normalized_term in normalized_category


def resolve_requested_categories(
    available_categories: pd.Series,
    requested_terms: list[str],
    source_path: Path,
) -> tuple[list[str], list[str]]:
    available = sorted(
        {
            str(category).strip()
            for category in available_categories.dropna().astype(str)
            if str(category).strip()
        }
    )

    matched: list[str] = []
    missing: list[str] = []
    seen: set[str] = set()
    for term in requested_terms:
        term_matches = [
            category for category in available if category_matches_term(category, term)
        ]
        if not term_matches:
            missing.append(term)
            continue
        for category in term_matches:
            if category not in seen:
                matched.append(category)
                seen.add(category)

    if matched:
        return matched, missing

    available_preview = ", ".join(available[:25])
    raise ValueError(
        f"No category terms from {source_path} matched expense transactions. "
        f"Available expense categories include: {available_preview}"
    )


def load_category_group_map(path: Path) -> dict[str, str]:
    if not path.exists():
        print(f"WARNING: {path} not found. Group columns will use {UNMAPPED_GROUP!r}.")
        return {}

    groups_df = read_csv(path)
    require_columns(groups_df, path, ["Category Name", "Group Name"])
    groups_df = groups_df[["Category Name", "Group Name"]].dropna().copy()
    groups_df["Category Name"] = groups_df["Category Name"].astype(str).str.strip()
    groups_df["Group Name"] = groups_df["Group Name"].astype(str).str.strip()
    groups_df = groups_df[
        (groups_df["Category Name"] != "") & (groups_df["Group Name"] != "")
    ]
    return dict(zip(groups_df["Category Name"], groups_df["Group Name"], strict=False))


def prepare_transactions(
    transactions_df: pd.DataFrame,
    transactions_path: Path,
    category_to_group: dict[str, str],
    optimizable_type: str,
    requested_terms: list[str],
    requested_terms_path: Path,
    *,
    lookback_months: int,
    expense_sign: str,
    include_hidden: bool,
) -> tuple[pd.DataFrame, pd.Timestamp, pd.Timestamp, list[str], list[str]]:
    require_columns(transactions_df, transactions_path, ["Date", "Amount", "Category"])

    prepared = transactions_df.copy()
    prepared["Date"] = parse_dates(prepared["Date"])
    prepared["Amount"] = prepared["Amount"].map(parse_amount)
    prepared = prepared.dropna(subset=["Date", "Amount"])
    prepared = prepared[prepared["Amount"] != 0].copy()

    if prepared.empty:
        raise ValueError(f"{transactions_path} did not contain dated non-zero amounts.")

    latest_date = prepared["Date"].max().normalize()
    cutoff_date = latest_date - pd.DateOffset(months=lookback_months)
    prepared = prepared[prepared["Date"] >= cutoff_date].copy()

    if not include_hidden and "Hide From Reports" in prepared.columns:
        hidden = prepared["Hide From Reports"].map(normalize_bool).fillna(False)
        prepared = prepared[~hidden].copy()

    if "Merchant" not in prepared.columns:
        prepared["Merchant"] = ""
    if "Plaid Name" in prepared.columns:
        merchant = clean_dimension(prepared["Merchant"])
        plaid_name = clean_dimension(prepared["Plaid Name"])
        prepared["Merchant"] = merchant.mask(merchant == BLANK_LABEL, plaid_name)
    else:
        prepared["Merchant"] = clean_dimension(prepared["Merchant"])

    for column in ["Account", "Category", "Tags"]:
        if column not in prepared.columns:
            prepared[column] = BLANK_LABEL
        prepared[column] = clean_dimension(prepared[column])

    prepared["Group"] = (
        prepared["Category"]
        .map(category_to_group)
        .fillna(UNMAPPED_GROUP)
        .astype(str)
        .str.strip()
        .replace("", UNMAPPED_GROUP)
    )

    expense_mask = (
        prepared["Amount"] > 0
        if expense_sign == "positive"
        else prepared["Amount"] < 0
    )
    prepared = prepared[expense_mask].copy()

    if prepared.empty:
        raise ValueError("No expense transactions remained after date/hidden filters.")

    if optimizable_type == "categories":
        matched_selections, missing_selections = resolve_requested_categories(
            prepared["Category"],
            requested_terms,
            requested_terms_path,
        )
        prepared = prepared[prepared["Category"].isin(matched_selections)].copy()
        prepared["Selection Type"] = "Category"
        prepared["Selection Match"] = prepared["Category"]
    else:
        matched_selections, missing_selections = resolve_requested_groups(
            prepared["Group"],
            requested_terms,
            requested_terms_path,
        )
        prepared = prepared[prepared["Group"].isin(matched_selections)].copy()
        prepared["Selection Type"] = "Group"
        prepared["Selection Match"] = prepared["Group"]

    if prepared.empty:
        raise ValueError(
            "No expense transactions remained after applying optimizable selections."
        )

    prepared["Flow Amount"] = prepared["Amount"].abs()
    prepared["Month"] = prepared["Date"].dt.to_period("M").astype(str)

    if "Transaction ID" not in prepared.columns:
        prepared["Transaction ID"] = prepared.index.astype(str)

    return (
        prepared,
        cutoff_date.normalize(),
        latest_date,
        matched_selections,
        missing_selections,
    )


def classify_cadence(
    dates: pd.Series,
    *,
    months_seen: int,
    active_window_months: int,
) -> tuple[str, float, float, float, float]:
    sorted_dates = dates.sort_values()
    gaps = sorted_dates.diff().dt.days.dropna()
    if gaps.empty:
        return "One-off", 0.0, 0.0, 0.0, 0.0

    median_gap = float(gaps.median())
    if median_gap <= 0:
        return "Same-day cluster", 0.0, median_gap, 0.0, 0.0

    gap_cv = float(gaps.std(ddof=0) / median_gap) if len(gaps) > 1 else 0.0
    cadence_bands = [
        (7.0, 2.0, "Weekly", 52.0),
        (14.0, 4.0, "Biweekly or Semimonthly", 26.0),
        (30.4375, 10.0, "Monthly", 12.0),
        (61.0, 14.0, "Every 2 Months", 6.0),
        (91.0, 21.0, "Quarterly", 4.0),
        (182.0, 35.0, "Semiannual", 2.0),
        (365.0, 70.0, "Annual", 1.0),
    ]

    best_name = "Irregular"
    best_fit = 0.0
    best_payments_per_year = safe_div(365.0, median_gap)
    for target_days, tolerance_days, name, payments_per_year in cadence_bands:
        fit = 1.0 - safe_div(abs(median_gap - target_days), tolerance_days)
        if fit > best_fit:
            best_name = name
            best_fit = fit
            best_payments_per_year = payments_per_year

    monthly_coverage = safe_div(months_seen, active_window_months)
    regularity_score = 1.0 - min(gap_cv, 1.0)

    if best_fit <= 0:
        if monthly_coverage >= 0.45:
            best_name = "Frequent Monthly Spend"
            best_fit = min(1.0, monthly_coverage)
            best_payments_per_year = 12.0
        else:
            best_name = f"Every {median_gap:.0f} Days"
            best_fit = 0.35 * regularity_score

    cadence_score = clamp(0.65 * best_fit + 0.35 * regularity_score)
    if best_name in {"Monthly", "Frequent Monthly Spend"}:
        cadence_score = max(cadence_score, clamp(monthly_coverage / 0.75) * 0.8)

    return best_name, best_payments_per_year, median_gap, gap_cv, cadence_score


def active_status(
    cadence: str,
    payments_per_year: float,
    days_since_last: int,
    recent_months: int,
) -> str:
    if payments_per_year > 0:
        expected_gap = safe_div(365.0, payments_per_year)
        if days_since_last <= expected_gap * 1.5 + 14:
            return "Active"
        if days_since_last <= expected_gap * 2.5 + 30:
            return "Watch"
        return "Possibly Ended"

    if cadence == "Frequent Monthly Spend" and days_since_last <= 45:
        return "Active"
    if days_since_last <= recent_months * 30:
        return "Watch"
    return "Possibly Ended"


def amount_stability_label(amount_cv: float, tolerance: float) -> str:
    if math.isnan(amount_cv):
        return "Unknown"
    if amount_cv <= tolerance:
        return "Stable"
    if amount_cv <= tolerance * 2:
        return "Slightly Variable"
    return "Variable"


def contains_any(text: str, terms: set[str]) -> bool:
    return any(term in text for term in terms)


def category_optimization_score(group: str, category: str) -> float:
    text = f"{group} {category}".casefold()
    if contains_any(text, SUBSCRIPTION_TERMS | NEGOTIATION_TERMS):
        return 1.0
    if contains_any(text, SHOPPING_TERMS):
        return 0.85
    if contains_any(text, LOW_CONTROL_TERMS):
        return 0.55
    return 0.75


def optimization_type(
    *,
    group: str,
    category: str,
    active: str,
    cadence: str,
    stability: str,
    estimated_annual: float,
    price_change: float,
    monthly_coverage: float,
) -> str:
    text = f"{group} {category}".casefold()
    if active == "Possibly Ended":
        return "Stale Recurring Charge"
    if price_change > 0.15:
        return "Price Increase Review"
    if (
        "Monthly" in cadence
        and stability in {"Stable", "Slightly Variable"}
        and contains_any(text, SUBSCRIPTION_TERMS)
    ):
        return "Subscription or Plan"
    if contains_any(text, NEGOTIATION_TERMS):
        return "Negotiation or Shopping"
    if stability == "Variable" and monthly_coverage >= 0.45:
        return "Usage or Habit Cap"
    if contains_any(text, SHOPPING_TERMS) and monthly_coverage >= 0.35:
        return "Habit-Like Spend"
    if estimated_annual >= 1200:
        return "High-Dollar Recurring Spend"
    return "Recurring Expense Review"


def priority_label(
    opportunity_score: float,
    estimated_annual: float,
    confidence: float,
) -> str:
    if confidence >= 65 and (opportunity_score >= 500 or estimated_annual >= 1200):
        return "High"
    if confidence >= 50 and (opportunity_score >= 150 or estimated_annual >= 300):
        return "Medium"
    return "Watch"


def recommendation(
    *,
    active: str,
    cadence: str,
    stability: str,
    estimated_annual: float,
    price_change: float,
    profile: str,
) -> str:
    if active == "Possibly Ended":
        return "Verify this recurring charge ended and clean up any unused service records."
    if price_change > 0.15:
        return "Recent amount is materially higher; check for price increases or plan changes."
    if profile == "Subscription or Plan":
        return "Cancel, downgrade, share, switch billing cadence, or compare annual pricing."
    if profile == "Negotiation or Shopping":
        return "Renegotiate, shop alternatives, adjust coverage, or right-size the plan."
    if profile in {"Usage or Habit Cap", "Habit-Like Spend"}:
        return "Set a monthly cap, reduce usage, compare alternatives, or consolidate purchases."
    if estimated_annual >= 1200:
        return "High-dollar recurring spend; review need, vendor, plan tier, and replacement options."
    if "Monthly" in cadence and stability == "Stable":
        return "Subscription-like recurring spend; verify it is still used and priced correctly."
    return "Recurring spend worth reviewing for cancellation, downgrade, negotiation, or alternatives."


def annualized_spend(
    *,
    cadence: str,
    payments_per_year: float,
    typical_amount: float,
    recent_monthly_average: float,
    trailing_12_amount: float,
) -> float:
    if payments_per_year > 0 and cadence != "Frequent Monthly Spend":
        cadence_annual = typical_amount * payments_per_year
        return max(cadence_annual, trailing_12_amount)
    return max(recent_monthly_average * 12.0, trailing_12_amount)


def analyze_candidate(
    candidate_df: pd.DataFrame,
    *,
    latest_date: pd.Timestamp,
    lookback_months: int,
    recent_months: int,
    amount_tolerance: float,
) -> dict[str, object]:
    candidate_df = candidate_df.sort_values("Date").copy()
    merchant = str(candidate_df["Merchant"].iloc[0])
    group = str(candidate_df["Group"].iloc[0])
    selection_type = str(candidate_df["Selection Type"].iloc[0])
    selection_match = str(candidate_df["Selection Match"].iloc[0])

    transaction_count = len(candidate_df)
    months_seen = candidate_df["Month"].nunique()
    first_date = candidate_df["Date"].min().normalize()
    last_date = candidate_df["Date"].max().normalize()
    active_window_months = min(lookback_months, month_count(first_date, latest_date))
    monthly_coverage = safe_div(months_seen, active_window_months)
    days_since_last = max(0, int((latest_date - last_date).days))

    cadence, payments_per_year, median_gap, gap_cv, cadence_score = classify_cadence(
        candidate_df["Date"],
        months_seen=months_seen,
        active_window_months=active_window_months,
    )

    amounts = candidate_df["Flow Amount"]
    typical_amount = float(amounts.median())
    mean_amount = float(amounts.mean())
    amount_std = float(amounts.std(ddof=0)) if len(amounts) > 1 else 0.0
    amount_cv = safe_div(amount_std, typical_amount)
    stability_score = 1.0 - min(
        safe_div(amount_cv, max(amount_tolerance * 2.5, 0.01)),
        1.0,
    )
    amount_signal = 0.60 + 0.40 * stability_score
    stability = amount_stability_label(amount_cv, amount_tolerance)

    first_amount = float(amounts.iloc[0])
    last_amount = float(amounts.iloc[-1])
    price_change = safe_div(last_amount - first_amount, abs(first_amount))

    recent_cutoff = latest_date - pd.DateOffset(months=recent_months)
    trailing_12_cutoff = latest_date - pd.DateOffset(months=12)
    recent_amount = float(
        candidate_df.loc[candidate_df["Date"] >= recent_cutoff, "Flow Amount"].sum()
    )
    trailing_12_amount = float(
        candidate_df.loc[candidate_df["Date"] >= trailing_12_cutoff, "Flow Amount"].sum()
    )
    total_lookback_amount = float(candidate_df["Flow Amount"].sum())
    recent_monthly_average = safe_div(recent_amount, max(1, recent_months))
    estimated_annual = annualized_spend(
        cadence=cadence,
        payments_per_year=payments_per_year,
        typical_amount=typical_amount,
        recent_monthly_average=recent_monthly_average,
        trailing_12_amount=trailing_12_amount,
    )

    status = active_status(cadence, payments_per_year, days_since_last, recent_months)
    recency_score = 1.0 - min(
        safe_div(days_since_last, max(1.0, recent_months * 30.4375)),
        1.0,
    )
    if payments_per_year > 0 and status == "Active":
        recency_score = max(recency_score, 0.65)
    elif status == "Watch":
        recency_score = max(recency_score, 0.35)

    frequency_score = min(1.0, safe_div(transaction_count, 12.0))
    if cadence == "Frequent Monthly Spend" and stability == "Variable":
        amount_signal = max(amount_signal, 0.75)

    top_category = top_values(candidate_df["Category"])
    primary_category = top_category.split(", ")[0]
    category_focus = safe_div(
        float(candidate_df["Category"].value_counts().iloc[0]),
        float(transaction_count),
    )
    category_score = category_optimization_score(group, primary_category)

    confidence = 100.0 * (
        0.45 * cadence_score
        + 0.25 * recency_score
        + 0.20 * frequency_score
        + 0.10 * amount_signal
    )
    confidence = round(confidence, 1)

    volatility_boost = 1.10 if stability == "Variable" and monthly_coverage >= 0.45 else 1.0
    price_boost = 1.10 if price_change > 0.15 else 1.0
    focus_boost = 0.90 + 0.10 * category_focus
    opportunity_score = (
        estimated_annual
        * (confidence / 100.0)
        * (0.65 + 0.35 * recency_score)
        * category_score
        * volatility_boost
        * price_boost
        * focus_boost
    )

    profile = optimization_type(
        group=group,
        category=primary_category,
        active=status,
        cadence=cadence,
        stability=stability,
        estimated_annual=estimated_annual,
        price_change=price_change,
        monthly_coverage=monthly_coverage,
    )

    row = {
        "Candidate ID": f"{merchant} | {selection_type} | {selection_match}",
        "Merchant": merchant,
        "Group": group,
        "Selection Type": selection_type,
        "Selection Match": selection_match,
        "Priority": priority_label(
            opportunity_score,
            estimated_annual,
            confidence,
        ),
        "Optimization Type": profile,
        "Top Category": top_category,
        "Accounts": top_values(candidate_df["Account"]),
        "Cadence": cadence,
        "Active Status": status,
        "Confidence": confidence,
        "Opportunity Score": round(opportunity_score, 2),
        "Transaction Count": transaction_count,
        "Months Seen": months_seen,
        "Monthly Coverage": round(monthly_coverage, 3),
        "First Transaction Date": first_date,
        "Last Transaction Date": last_date,
        "Days Since Last Transaction": days_since_last,
        "Median Gap Days": round(median_gap, 1),
        "Gap Coefficient Variation": round(gap_cv, 3),
        "Amount Stability": stability,
        "Amount Coefficient Variation": round(amount_cv, 3),
        "Category Focus": round(category_focus, 3),
        "Optimizable Selection Score": round(category_score, 2),
        "Typical Amount": round(typical_amount, 2),
        "Average Amount": round(mean_amount, 2),
        "Last Amount": round(last_amount, 2),
        "Price Change Since First Seen": round(price_change, 3),
        "Recent Monthly Average": round(recent_monthly_average, 2),
        "Trailing 12 Month Spend": round(trailing_12_amount, 2),
        "Lookback Total Spend": round(total_lookback_amount, 2),
        "Estimated Annual Spend": round(estimated_annual, 2),
        "Potential Annual Savings 10%": round(estimated_annual * 0.10, 2),
        "Potential Annual Savings 25%": round(estimated_annual * 0.25, 2),
        "Potential Annual Savings 50%": round(estimated_annual * 0.50, 2),
        "Potential Annual Savings 100%": round(estimated_annual, 2),
    }
    row["Recommendation"] = recommendation(
        active=status,
        cadence=cadence,
        stability=stability,
        estimated_annual=estimated_annual,
        price_change=price_change,
        profile=profile,
    )
    return row


def build_candidates(
    df: pd.DataFrame,
    *,
    latest_date: pd.Timestamp,
    lookback_months: int,
    recent_months: int,
    min_occurrences: int,
    min_months: int,
    min_confidence: float,
    min_annualized_spend: float,
    amount_tolerance: float,
    top: int,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []

    for (_merchant, _selection_type, _selection_match), candidate_df in df.groupby(
        ["Merchant", "Selection Type", "Selection Match"],
        dropna=False,
        sort=False,
    ):
        if len(candidate_df) < min_occurrences:
            continue
        if candidate_df["Month"].nunique() < min_months:
            continue

        row = analyze_candidate(
            candidate_df,
            latest_date=latest_date,
            lookback_months=lookback_months,
            recent_months=recent_months,
            amount_tolerance=amount_tolerance,
        )
        if float(row["Confidence"]) < min_confidence:
            continue
        if float(row["Estimated Annual Spend"]) < min_annualized_spend:
            continue

        rows.append(row)

    if not rows:
        return pd.DataFrame(columns=ACTION_COLUMNS)

    candidates = pd.DataFrame(rows)
    priority_rank = {"High": 0, "Medium": 1, "Watch": 2}
    candidates["__priority_rank"] = candidates["Priority"].map(priority_rank).fillna(9)
    candidates = candidates.sort_values(
        ["__priority_rank", "Opportunity Score", "Estimated Annual Spend", "Merchant"],
        ascending=[True, False, False, True],
        kind="stable",
    ).drop(columns=["__priority_rank"])

    return candidates.head(top).reset_index(drop=True)


def action_plan(candidates: pd.DataFrame) -> pd.DataFrame:
    if candidates.empty:
        return pd.DataFrame(columns=ACTION_COLUMNS)
    return candidates[[column for column in ACTION_COLUMNS if column in candidates.columns]]


def group_summary(candidates: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Group",
        "Candidates",
        "High Priority",
        "Estimated Annual Spend",
        "Potential Annual Savings 25%",
        "Opportunity Score",
        "Average Confidence",
        "Top Merchants",
    ]
    if candidates.empty:
        return pd.DataFrame(columns=columns)

    rows: list[dict[str, object]] = []
    for group, group_df in candidates.groupby("Group", sort=False):
        sorted_group = group_df.sort_values(
            ["Opportunity Score", "Estimated Annual Spend"],
            ascending=[False, False],
            kind="stable",
        )
        rows.append(
            {
                "Group": group,
                "Candidates": len(group_df),
                "High Priority": int((group_df["Priority"] == "High").sum()),
                "Estimated Annual Spend": float(group_df["Estimated Annual Spend"].sum()),
                "Potential Annual Savings 25%": float(
                    group_df["Potential Annual Savings 25%"].sum()
                ),
                "Opportunity Score": float(group_df["Opportunity Score"].sum()),
                "Average Confidence": float(group_df["Confidence"].mean()),
                "Top Merchants": ", ".join(sorted_group["Merchant"].head(5).astype(str)),
            }
        )

    summary = pd.DataFrame(rows)
    return summary.sort_values(
        ["Opportunity Score", "Estimated Annual Spend", "Group"],
        ascending=[False, False, True],
        kind="stable",
    ).reset_index(drop=True)


def category_summary(
    df: pd.DataFrame,
    candidates: pd.DataFrame,
    *,
    latest_date: pd.Timestamp,
    recent_months: int,
) -> pd.DataFrame:
    columns = [
        "Group",
        "Category",
        "Merchants",
        "Transaction Count",
        "Months Seen",
        "Recent Monthly Average",
        "Trailing 12 Month Spend",
        "Lookback Total Spend",
        "Last Transaction Date",
    ]
    if candidates.empty:
        return pd.DataFrame(columns=columns)

    candidate_ids = set(candidates["Candidate ID"].astype(str))
    scoped = df[df["Candidate ID"].isin(candidate_ids)].copy()
    if scoped.empty:
        return pd.DataFrame(columns=columns)

    recent_cutoff = latest_date - pd.DateOffset(months=recent_months)
    trailing_12_cutoff = latest_date - pd.DateOffset(months=12)
    rows: list[dict[str, object]] = []

    for (group, category), category_df in scoped.groupby(
        ["Group", "Category"],
        dropna=False,
        sort=False,
    ):
        recent_amount = float(
            category_df.loc[category_df["Date"] >= recent_cutoff, "Flow Amount"].sum()
        )
        trailing_12_amount = float(
            category_df.loc[
                category_df["Date"] >= trailing_12_cutoff,
                "Flow Amount",
            ].sum()
        )
        rows.append(
            {
                "Group": group,
                "Category": category,
                "Merchants": category_df["Merchant"].nunique(),
                "Transaction Count": len(category_df),
                "Months Seen": category_df["Month"].nunique(),
                "Recent Monthly Average": safe_div(recent_amount, max(1, recent_months)),
                "Trailing 12 Month Spend": trailing_12_amount,
                "Lookback Total Spend": float(category_df["Flow Amount"].sum()),
                "Last Transaction Date": category_df["Date"].max().normalize(),
            }
        )

    summary = pd.DataFrame(rows)
    return summary.sort_values(
        ["Trailing 12 Month Spend", "Lookback Total Spend", "Group", "Category"],
        ascending=[False, False, True, True],
        kind="stable",
    ).reset_index(drop=True)


def monthly_trend(df: pd.DataFrame, candidates: pd.DataFrame) -> pd.DataFrame:
    if candidates.empty:
        return pd.DataFrame(
            columns=[
                "Selection Type",
                "Selection Match",
                "Group",
                "Merchant",
                TOTAL_LABEL,
            ]
        )

    candidate_ids = set(candidates["Candidate ID"].astype(str))
    scoped = df[df["Candidate ID"].isin(candidate_ids)].copy()
    if scoped.empty:
        return pd.DataFrame(
            columns=[
                "Selection Type",
                "Selection Match",
                "Group",
                "Merchant",
                TOTAL_LABEL,
            ]
        )

    trend = pd.pivot_table(
        scoped,
        index=["Candidate ID", "Selection Type", "Selection Match", "Group", "Merchant"],
        columns="Month",
        values="Flow Amount",
        aggfunc="sum",
        fill_value=0.0,
    )
    month_columns = sorted(str(column) for column in trend.columns)
    trend = trend[month_columns]
    trend[TOTAL_LABEL] = trend.sum(axis=1)
    trend = trend.reset_index()
    trend.columns = [str(column) for column in trend.columns]

    totals = (
        candidates[["Candidate ID", "Opportunity Score"]]
        .drop_duplicates()
        .copy()
    )
    trend = trend.merge(totals, how="left", on="Candidate ID")
    trend = trend.sort_values(
        ["Opportunity Score", TOTAL_LABEL, "Selection Match", "Merchant"],
        ascending=[False, False, True, True],
        kind="stable",
    )
    return trend.drop(columns=["Candidate ID", "Opportunity Score"])


def transaction_detail(df: pd.DataFrame, candidates: pd.DataFrame) -> pd.DataFrame:
    if candidates.empty:
        return pd.DataFrame(
            columns=[
                "Candidate ID",
                "Selection Type",
                "Selection Match",
                "Transaction ID",
                "Date",
                "Merchant",
                "Account",
                "Category",
                "Group",
                "Amount",
                "Flow Amount",
                "Tags",
                "Notes",
            ]
        )

    candidate_ids = set(candidates["Candidate ID"].astype(str))
    detail = df[df["Candidate ID"].isin(candidate_ids)].copy()
    preferred = [
        "Candidate ID",
        "Selection Type",
        "Selection Match",
        "Transaction ID",
        "Date",
        "Merchant",
        "Account",
        "Category",
        "Group",
        "Amount",
        "Flow Amount",
        "Tags",
        "Notes",
        "Hide From Reports",
        "Needs Review",
    ]
    columns = [column for column in preferred if column in detail.columns]
    detail = detail[columns]
    return detail.sort_values(
        ["Selection Type", "Selection Match", "Merchant", "Date", "Flow Amount"],
        ascending=[True, True, True, False, False],
        kind="stable",
    )


def summary_sheet(
    *,
    source_path: Path,
    groups_path: Path,
    optimizable_type: str,
    optimizable_terms_path: Path,
    output_path: Path,
    row_count: int,
    analyzed_count: int,
    candidate_count: int,
    cutoff_date: pd.Timestamp,
    latest_date: pd.Timestamp,
    matched_selections: list[str],
    missing_selections: list[str],
    args: argparse.Namespace,
) -> pd.DataFrame:
    rows = [
        ("Source transactions", str(source_path)),
        ("Category group map", str(groups_path)),
        ("Optimizable selection type", optimizable_type),
        ("Optimizable selection file", str(optimizable_terms_path)),
        ("Output workbook", str(output_path)),
        ("Rows read", row_count),
        ("Expense rows analyzed after all filters", analyzed_count),
        ("Expense candidates included", candidate_count),
        ("Analysis start date", cutoff_date.date().isoformat()),
        ("Latest transaction date", latest_date.date().isoformat()),
        ("Lookback months", min(max(args.lookback_months, 1), 36)),
        ("Recent months", max(args.recent_months, 1)),
        ("Minimum occurrences", max(args.min_occurrences, 1)),
        ("Minimum months seen", max(args.min_months, 1)),
        ("Minimum confidence", args.min_confidence),
        ("Minimum annualized spend", args.min_annualized_spend),
        ("Amount tolerance", args.amount_tolerance),
        ("Expense sign", args.expense_sign),
        ("Included hidden transactions", args.include_hidden),
        ("Matched optimizable selections", ", ".join(matched_selections)),
        (
            "Missing requested selections",
            ", ".join(missing_selections) if missing_selections else "",
        ),
        (
            "Method",
            "Filters to expenses in the selected groups or category wildcard "
            "matches, groups by merchant and selected dimension, then scores "
            "cadence, recency, frequency, amount variability, "
            "category focus, price movement, and optimization category.",
        ),
        (
            "Goal",
            "Find recurring or habit-like expense streams that can be cancelled, "
            "downgraded, negotiated, capped, consolidated, or moved to cheaper alternatives.",
        ),
        (
            "Caution",
            "This is a prioritization report, not a financial recommendation. "
            "Review transaction details before changing services.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Setting", "Value"])


def build_sheets(
    df: pd.DataFrame,
    candidates: pd.DataFrame,
    *,
    source_path: Path,
    groups_path: Path,
    optimizable_type: str,
    optimizable_terms_path: Path,
    output_path: Path,
    row_count: int,
    cutoff_date: pd.Timestamp,
    latest_date: pd.Timestamp,
    matched_selections: list[str],
    missing_selections: list[str],
    args: argparse.Namespace,
) -> dict[str, pd.DataFrame]:
    df = df.copy()
    df["Candidate ID"] = (
        df["Merchant"] + " | " + df["Selection Type"] + " | " + df["Selection Match"]
    )

    return {
        "Summary": summary_sheet(
            source_path=source_path,
            groups_path=groups_path,
            optimizable_type=optimizable_type,
            optimizable_terms_path=optimizable_terms_path,
            output_path=output_path,
            row_count=row_count,
            analyzed_count=len(df),
            candidate_count=len(candidates),
            cutoff_date=cutoff_date,
            latest_date=latest_date,
            matched_selections=matched_selections,
            missing_selections=missing_selections,
            args=args,
        ),
        "Action Plan": action_plan(candidates),
        "Expense Opportunities": candidates,
        "Group Summary": group_summary(candidates),
        "Category Summary": category_summary(
            df,
            candidates,
            latest_date=latest_date,
            recent_months=args.recent_months,
        ),
        "Monthly Trend": monthly_trend(df, candidates),
        "Transactions": transaction_detail(df, candidates),
    }


def is_locked_file_error(error: OSError) -> bool:
    return isinstance(error, PermissionError) or getattr(error, "winerror", None) in {
        32,
        33,
    }


def format_workbook(writer: pd.ExcelWriter) -> None:
    for ws in writer.sheets.values():
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = {cell.column: str(cell.value or "").strip() for cell in ws[1]}
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                header = headers.get(cell.column, "")
                if "Date" in header:
                    cell.number_format = DATE_FORMAT
                elif (
                    "Amount" in header
                    or "Spend" in header
                    or "Savings" in header
                    or "Average" in header
                    or header == TOTAL_LABEL
                    or header.startswith("20")
                ):
                    cell.number_format = AMOUNT_FORMAT
                elif (
                    "Count" in header
                    or "Months Seen" in header
                    or "Days Since" in header
                    or header == "Merchants"
                    or header == "Candidates"
                    or header == "High Priority"
                ):
                    cell.number_format = INTEGER_FORMAT
                elif (
                    "Variation" in header
                    or "Change" in header
                    or "Coverage" in header
                    or "Focus" in header
                ):
                    cell.number_format = PERCENT_FORMAT
                elif "Confidence" in header or "Score" in header or "Gap Days" in header:
                    cell.number_format = SCORE_FORMAT

        for column_cells in ws.columns:
            header = str(column_cells[0].value or "")
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_length + 2, 10), 55)
            if header in {"Recommendation", "Value", "Top Merchants"}:
                width = 70
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_workbook(output_path: Path, sheets: dict[str, pd.DataFrame]) -> bool:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, sheet_df in sheets.items():
                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
            format_workbook(writer)
    except OSError as e:
        if not is_locked_file_error(e):
            raise
        print(f"WARNING: Could not write {output_path}.")
        print("         It may be open in Excel. Close it and run the report again.")
        return False

    return True


def normalized_args(args: argparse.Namespace) -> argparse.Namespace:
    args.lookback_months = min(max(args.lookback_months, 1), 36)
    args.recent_months = max(args.recent_months, 1)
    args.min_occurrences = max(args.min_occurrences, 1)
    args.min_months = max(args.min_months, 1)
    args.amount_tolerance = max(args.amount_tolerance, 0.01)
    args.top = max(args.top, 1)
    return args


def main() -> None:
    args = normalized_args(parse_args())
    optimizable_terms_path = (
        args.optimizable_categories
        if args.optimizable_type == "categories"
        else args.optimizable_groups
    )
    optimizable_label = (
        "category search term"
        if args.optimizable_type == "categories"
        else "category group name"
    )
    requested_terms = load_optimizable_terms(optimizable_terms_path, optimizable_label)
    category_to_group = load_category_group_map(args.groups)
    transactions_df = read_csv(args.transactions)
    source_count = len(transactions_df)
    (
        prepared_df,
        cutoff_date,
        latest_date,
        matched_selections,
        missing_selections,
    ) = prepare_transactions(
        transactions_df,
        args.transactions,
        category_to_group,
        args.optimizable_type,
        requested_terms,
        optimizable_terms_path,
        lookback_months=args.lookback_months,
        expense_sign=args.expense_sign,
        include_hidden=args.include_hidden,
    )
    candidates = build_candidates(
        prepared_df,
        latest_date=latest_date,
        lookback_months=args.lookback_months,
        recent_months=args.recent_months,
        min_occurrences=args.min_occurrences,
        min_months=args.min_months,
        min_confidence=args.min_confidence,
        min_annualized_spend=args.min_annualized_spend,
        amount_tolerance=args.amount_tolerance,
        top=args.top,
    )
    sheets = build_sheets(
        prepared_df,
        candidates,
        source_path=args.transactions,
        groups_path=args.groups,
        optimizable_type=args.optimizable_type,
        optimizable_terms_path=optimizable_terms_path,
        output_path=args.output,
        row_count=source_count,
        cutoff_date=cutoff_date,
        latest_date=latest_date,
        matched_selections=matched_selections,
        missing_selections=missing_selections,
        args=args,
    )

    print(f"Read {source_count} transaction rows from {args.transactions}")
    print(
        f"Analyzed {len(prepared_df)} expense rows from {cutoff_date.date()} "
        f"through {latest_date.date()}"
    )
    print(f"Loaded {len(category_to_group)} category-to-group mappings from {args.groups}")
    print(
        f"Loaded {len(requested_terms)} optimizable {optimizable_label}s from "
        f"{optimizable_terms_path}"
    )
    print(f"Selection type: {args.optimizable_type}")
    print(f"Matched selections: {', '.join(matched_selections)}")
    if missing_selections:
        print(
            "WARNING: Requested selections not found in expense data: "
            f"{', '.join(missing_selections)}"
        )
    print(f"Recurring expense candidates: {len(candidates)}")
    if not write_workbook(args.output, sheets):
        return

    print(f"Saved recurring expense optimization workbook to {args.output}")


if __name__ == "__main__":
    main()
